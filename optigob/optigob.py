from matplotlib import pyplot as plt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from configuration.keys import *

from resource_manager.database_manager import DatabaseManager

from .systems.cattle_agriculture import CattleAgriculture
from .systems.forestry import Forestry
from .systems.non_cattle_agriculture import NonCattleAgriculture
from .systems.organic_soils import OrganicSoils
from .systems.ad_emissions import AnaerobicDigestion
from .utils import add_two_lists, transform_to_c02e


class Optigob:

    def __init__(self, json_config, db_file_path):
        self.baseline_year = json_config[BASELINE_YEAR]
        self.target_year = json_config[TARGET_YEAR]
        self.fields = []
        self.db_manager = DatabaseManager(db_file_path)

        if FORESTRY in json_config:
            self.fields.append(Forestry(json_config[FORESTRY]))

        if NON_CATTLE_AGRICULTURE in json_config:
            self.fields.append(NonCattleAgriculture(json_config[NON_CATTLE_AGRICULTURE]))

        if CATTLE_AGRICULTURE in json_config:
            self.fields.append((CattleAgriculture(json_config[CATTLE_AGRICULTURE])))

        if ORGANIC_SOILS in json_config:
            self.fields.append(OrganicSoils(json_config[ORGANIC_SOILS]))

        if AD_EMISSIONS in json_config:
            self.fields.append(AnaerobicDigestion(json_config[AD_EMISSIONS]))

        for fi in self.fields:
            fi.load_data(self.db_manager)

        self.apply_scalers()

    def apply_scalers(self):
        scalers = self.db_manager.get_scalers()
        for fi in self.fields:
            for system in fi.systems:
                if system.name in scalers.keys():
                    for i in range(len(scalers["Year"])):
                        system.update_by_scaler(scaler=scalers[system.name][i],
                                                baseline_year=self.baseline_year,
                                                target_year=scalers["Year"][i])

    def run(self):
        nca = None
        for fi in self.fields:
            fi.run(self.baseline_year, self.target_year, self.db_manager)

            if isinstance(fi, NonCattleAgriculture):
                nca = fi

        for fi in self.fields:
            if isinstance(fi, CattleAgriculture):
                fi.run_cattle_systems(self.baseline_year, self.target_year, self.db_manager, nca)

        self.area_balancing()

    def area_balancing(self):
        # this function only handles balancing of area within different fields
        # organic soil rewetting balancing in systems.organic_soils.OrganicSoilSystem.run()
        # croplands/no-croplands balancing in systems.non_cattle_agriculture.NonCattleAgriculture.run()

        # beef.area_beef    =>                          => afforestation.area
        # dairy.area_dairy  => spared_sheep_cattle.area => ad.area
        # dairy.area_beef   =>                          => additional_ad.area
        # sheep.area        =>                          => willow_ad.area
        if self.get_field(CATTLE_AGRICULTURE) is not None:
            self.balance_spared_sheep_cattle_area()

        # organic_soil_under_grass.drained_area => afforestation.organic_soil_area
        if self.get_field(ORGANIC_SOILS) is not None and self.get_field(ORGANIC_SOILS).get_system(ORGANIC_SOILS_ORGANIC_SOIL_UNDER_GRASS) is not None and self.get_field(FORESTRY) is not None and self.get_field(FORESTRY).get_system(FORESTRY_AFFORESTATION) is not None:
            self.balance_afforestation_organic_soils()

    def balance_spared_sheep_cattle_area(self):
        assert self.get_field(CATTLE_AGRICULTURE) is not None
        sheep_included = self.get_field(NON_CATTLE_AGRICULTURE) is not None and self.get_field(NON_CATTLE_AGRICULTURE).get_system(NON_CATTLE_AGRICULTURE_SHEEP) is not None
        afforestation_included = self.get_field(FORESTRY) is not None and self.get_field(FORESTRY).get_system(FORESTRY_AFFORESTATION) is not None
        ad_included = self.get_field(AD_EMISSIONS) is not None

        dairy = self.get_field(CATTLE_AGRICULTURE).get_system(CATTLE_AGRICULTURE_DAIRY)
        beef = self.get_field(CATTLE_AGRICULTURE).get_system(CATTLE_AGRICULTURE_BEEF)
        sheep, afforestation, ad_emissions = None, None, None

        if sheep_included:
            sheep = self.get_field(NON_CATTLE_AGRICULTURE).get_system(NON_CATTLE_AGRICULTURE_SHEEP)
        if afforestation_included:
            afforestation = self.get_field(FORESTRY).get_system(FORESTRY_AFFORESTATION)
        if ad_included:
            ad_emissions = self.get_field(AD_EMISSIONS).get_system(AD_EMISSIONS)

        for i in range(self.target_year - self.baseline_year + 1):
            diff = (dairy.time_series[DAIRY_AREA][0] - dairy.time_series[DAIRY_AREA][i]
                     + dairy.time_series[BEEF_AREA][0] - dairy.time_series[BEEF_AREA][i]
                     + beef.time_series[BEEF_AREA][0] - beef.time_series[BEEF_AREA][i])
            if sheep_included:
                diff += sheep.time_series[AREA][0] - sheep.time_series[AREA][i]
            if afforestation_included:
                diff += afforestation.time_series[AREA][0] - afforestation.time_series[AREA][i]
            if ad_included:
                diff += (ad_emissions.time_series[AREA][0] - ad_emissions.time_series[AREA][i]
                         + ad_emissions.time_series[AD_ADDITIONAL_AREA][0] - ad_emissions.time_series[AD_ADDITIONAL_AREA][i]
                         + ad_emissions.time_series[AD_WILLOW_AREA][0] - ad_emissions.time_series[AD_WILLOW_AREA][i])

            self.get_field(CATTLE_AGRICULTURE).get_system(CATTLE_AGRICULTURE_SPARED_AREA).time_series[AREA][i] += diff

    def balance_afforestation_organic_soils(self):
        afforestation = self.get_field(FORESTRY).get_system(FORESTRY_AFFORESTATION)
        organic_soils_under_grass = self.get_field(ORGANIC_SOILS).get_system(ORGANIC_SOILS_ORGANIC_SOIL_UNDER_GRASS)

        for i in range(self.target_year - self.baseline_year + 1):
            diff = afforestation.time_series[AFFORESTATION_ORGANIC_SOIL_AREA][0] - afforestation.time_series[AFFORESTATION_ORGANIC_SOIL_AREA][i]
            new_area = organic_soils_under_grass.time_series[DRAINED + "_" + AREA][i] + diff
            self.get_field(ORGANIC_SOILS).get_system(ORGANIC_SOILS_ORGANIC_SOIL_UNDER_GRASS).area_balance(i, new_area, DRAINED)

    def get_evaluation(self, parameter):
        output_list = []
        time_span = self.target_year - self.baseline_year + 1

        for f in self.fields:
            field_list = None
            if parameter == CO2E:
                field_list = f.get_co2e(time_span)
            elif parameter == AREA:
                field_list = f.get_area(time_span)
            elif parameter == PROTEIN:
                field_list = f.get_protein(time_span)
            elif parameter == BIO_ENERGY:
                field_list = f.get_bio_energy(time_span)
            elif parameter == HWP:
                field_list = f.get_hwp(time_span)
            elif parameter == SUBSTITUTION:
                field_list = f.get_substitution(time_span)
            elif parameter == BIODIVERSITY:
                field_list = f.get_biodiversity(time_span)

            if not field_list is None:
                output_list.extend(field_list)

        i = 0
        while i < len(output_list):
            (label, value) = output_list[i]
            if sum(value) == 0:
                rm = True
                for v in value:
                    if v != 0.0:
                        rm = False
                if rm:
                    output_list.pop(i)
                else:
                    i += 1
            else:
                i += 1

        if parameter == CO2E:
            co2e, co2e_split_gas, total_ch4 = self.get_net_zero_calculations()
            output_list.append(("net_zero_co2e", co2e))
            output_list.append(("net_zero_split_gas_co2/n2o", co2e_split_gas))
            output_list.append(("net_zero_split_gas_ch4", total_ch4))

        return output_list

    def get_net_zero_calculations(self):
        time_span = self.target_year - self.baseline_year + 1
        total_co2, total_n2o, total_ch4 = [], [], []
        for f in self.fields:
            (co2, n2o, ch4) = f.get_net_zero(time_span=time_span)
            total_co2 = add_two_lists(total_co2, co2)
            total_n2o = add_two_lists(total_n2o, n2o)
            total_ch4 = add_two_lists(total_ch4, ch4)

        co2e = []
        co2e_split_gas = []
        for i in range(time_span):
            co2e.append(transform_to_c02e(co2=total_co2[i], n2o=total_n2o[i], ch4=total_ch4[i]))
            co2e_split_gas.append(transform_to_c02e(co2=total_co2[i], n2o=total_n2o[i], ch4=0))

        return co2e, co2e_split_gas, total_ch4

    def get_field(self, name):
        for f in self.fields:
            if f.name == name:
                return f
        return None

    def visualise(self, parameter, systems):
        x = range(self.baseline_year, self.target_year + 1)

        totals = []
        for i in x:
            totals.append(0)

        vis_systems = []
        for f in self.fields:
            if f.name in systems:
                vis_systems.extend(f.systems)
            else:
                for system in f.systems:
                    if system.name in systems:
                        vis_systems.append(system)

        for s in vis_systems:
            results = s.get(parameter)
            for (label, y) in results:
                plt.plot(x, y, label=label)
                for i in range(len(y)):
                    totals[i] += y[i]

        plt.plot(x, totals, label="total")

        plt.xlabel("Timeline")
        plt.ylabel("parameter")
        plt.legend()
        plt.show()

    def export_time_series(self):
        wb = Workbook()
        # Remove default sheet (optional, cleaner)
        wb.remove(wb.active)

        for f in self.fields:
            sheet = wb.create_sheet(title=f.name)

            data = [["System"], ["Parameter"], ["Year"]]
            for i in range(self.baseline_year, self.target_year + 1):
                data.append([str(i)])

            for s in f.systems:
                for p in s.time_series.keys():
                    data[0].append(s.name)
                    data[1].append(p)
                    data[2].append("")
                    for i in range(self.baseline_year, self.target_year + 1):
                        idx = i - self.baseline_year
                        data[3+idx].append(s.time_series[p][idx])

            # ---- write data to sheet ----
            for row in data:
                sheet.append(row)

            # ---- create dynamic table ----
            max_row = sheet.max_row
            max_col = sheet.max_column

            table_ref = f"A1:{get_column_letter(max_col)}{max_row}"

            table = Table(displayName=f"Table_{f.name}", ref=table_ref)

            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
                showColumnStripes=False
            )

            sheet.add_table(table)

        # Save to memory
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
